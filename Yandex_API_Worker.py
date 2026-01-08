import requests
from argparse import ArgumentParser
from prettytable import PrettyTable
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ORG_ID = "bpf9jinr4ulcogpcdv8i"
TOKEN = "y0__xDyhLZKGLCXOyCj6Lv0FDCi5JKLCN0MHANdHcxxWaB_WYMkwJ4YSJna"


def get_issues(queue):
    headers = {
        "Authorization": f"OAuth {TOKEN}",
        "X-Cloud-Org-ID": ORG_ID,
        "Content-Type": "application/json"
    }

    body = {
        "filter": {
            "queue": queue
        }
    }

    # Используем максимальный размер страницы в URL
    url = "https://api.tracker.yandex.net/v2/issues/_search?perPage=1000"
    response = requests.post(url, headers=headers, json=body)
    response.raise_for_status()

    return response.json()


def print_issues_table(issues):
    table = PrettyTable()
    table.field_names = ["KEY", "STATUS", "ASSIGNEE", "SUMMARY"]
    table.align = "l"

    for issue in issues:
        key = issue["key"]
        status = issue.get("status", {}).get("name", "No status")
        assignee = issue.get("assignee", {}).get("login", "Unassigned")
        summary = issue["summary"]

        table.add_row([key, status, assignee, summary])

    print(table)


def calculate_stats(issues):
    stats = {}
    for issue in issues:
        status = issue.get("status", {}).get("name", "No status")
        stats[status] = stats.get(status, 0) + 1
    return stats


def save_issues_to_csv(issues, queue_name):
    with open(f"{queue_name}_issues.csv", "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["KEY", "STATUS", "ASSIGNEE", "SUMMARY"])

        for issue in issues:
            key = issue["key"]
            status = issue.get("status", {}).get("name", "No status")
            assignee = issue.get("assignee", {}).get("login", "Unassigned")
            summary = issue["summary"]

            writer.writerow([key, status, assignee, summary])


def save_stats_to_csv(stats, queue_name):
    with open(f"{queue_name}_stats.csv", "w", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Status", "Count"])
        for status, count in stats.items():
            writer.writerow([status, count])


def save_issues_to_xlsx(issues, stats, queue_name):
    wb = Workbook()

    # Лист 1: Таблица задач
    ws_issues = wb.active
    ws_issues.title = "Issues"

    # Стиль для заголовка
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Заголовки
    headers = ["KEY", "STATUS", "ASSIGNEE", "SUMMARY"]
    ws_issues.append(headers)

    for cell in ws_issues[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные задач
    for issue in issues:
        key = issue["key"]
        status = issue.get("status", {}).get("name", "No status")
        assignee = issue.get("assignee", {}).get("login", "Unassigned")
        summary = issue["summary"]

        ws_issues.append([key, status, assignee, summary])

    # Автоматическое выравнивание ширины колонок
    ws_issues.column_dimensions['A'].width = 12
    ws_issues.column_dimensions['B'].width = 15
    ws_issues.column_dimensions['C'].width = 15
    ws_issues.column_dimensions['D'].width = 40

    # Лист 2: Статистика
    ws_stats = wb.create_sheet("Statistics")

    stat_headers = ["Status", "Count"]
    ws_stats.append(stat_headers)

    for cell in ws_stats[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for status, count in stats.items():
        ws_stats.append([status, count])

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 12

    # Сохранение файла
    wb.save(f"{queue_name}_report.xlsx")


def main():
    parser = ArgumentParser(description="Yandex Tracker Queue Analyzer")
    parser.add_argument("--queue", required=True, help="Queue name to analyze")
    parser.add_argument("--csv", action="store_true", help="Save stats to CSV file")
    parser.add_argument("--xlsx", action="store_true", help="Save to Excel file")
    args = parser.parse_args()

    try:
        issues = get_issues(args.queue)
        if not issues:
            print(f"No issues found in queue {args.queue}")
            return

        print_issues_table(issues)

        stats = calculate_stats(issues)
        print("\nStatus statistics:")
        for status, count in stats.items():
            print(f"{status}: {count} issues")

        if args.csv:
            save_issues_to_csv(issues, args.queue)
            save_stats_to_csv(stats, args.queue)
            print(f"\nData saved to {args.queue}_issues.csv")
            print(f"Statistics saved to {args.queue}_stats.csv")

        if args.xlsx:
            save_issues_to_xlsx(issues, stats, args.queue)
            print(f"\nReport saved to {args.queue}_report.xlsx")

    except requests.exceptions.HTTPError as e:
        print(f"API request failed: {e.response.status_code} {e.response.reason}")
    except Exception as e:
        print(f"Error: {str(e)}")


if __name__ == "__main__":
    main()